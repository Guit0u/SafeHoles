from PySide2 import QtCore, QtGui, QtWidgets
import numpy as np
import sys
import re
import openpyxl
import tkinter as tk
from tkinter.filedialog import askopenfilename
from openpyxl import Workbook
import glob
import os
import pathlib


##Fenetre utilisateur

def parse_float(str_value):
    """
    This fuction converts a string to float just like the built-in
    float() function. In addition to "normal" numbers it also handles
    numbers such as 1.2D3 (equivalent to 1.2E3)
    """
    try:
        return float(str_value)
    except ValueError:
        return float(str_value.lower().replace("d", "e"))


class MaFenetre(QtWidgets.QMainWindow):

    def __init__(self, parent=None):
        super().__init__()

        ##variables globales à la con
        self.chosenFile = ''
        self.points = []
        self.infopoints = []
        self.compteur = 0
        self.Fanuc = False
        self.Gxx = 'G'
        self.Mx = 'M0'
        self.Sxx = 'S0'
        self.Fxx = 'F0'
        self.Rxx = 'R0'
        self.Qxx = 'Q0'
        self.Securite = 'Z'
        self.Profondeur = 'Z-0'
        self.Diametre= ''

        self.__bravo = QtWidgets.QLabel("File created")
        layout7=QtWidgets.QGridLayout()
        layout7.addWidget(self.__bravo)
        self.widget7=QtWidgets.QWidget()
        self.widget7.setLayout(layout7)


        ##layout1 : rentrer le fichier et le type de sortie
        # self.boutonLire = QtWidgets.QPushButton("entrata")
        self.__champTexte = QtWidgets.QLineEdit("")
        self.__champTexte.setPlaceholderText("esempio.igs")
        self.__labelText = QtWidgets.QLabel("Inserisci il nome del file")
        self.__champSecurite = QtWidgets.QLineEdit("")
        self.__champSecurite.setPlaceholderText("100")
        self.__unite = QtWidgets.QLabel("Sicurezza (millimetri)")
        self.__error1 = QtWidgets.QLabel()
        self.__error11 = QtWidgets.QLabel()
        self.__error2 = QtWidgets.QLabel()
        self.__nul = QtWidgets.QLabel()
        self.__nul2 = QtWidgets.QLabel()

        self.__buttonChoix = QtWidgets.QPushButton("Select File")
        self.__buttonChoixLabel = QtWidgets.QLabel("File : ")

        layout1 = QtWidgets.QGridLayout()
       # layout1.addWidget(self.__champTexte, 1, 1)
        # layout1.addWidget(self.boutonLire, 4,1)
        # layout1.addWidget(self.__champSecurite,1,1)
        # layout1.addWidget(self.__unite,1,0)
        self.__nul2.setText('Scegli il tipo di file di output')
        layout1.addWidget(self.__nul, 8, 1)
        layout1.addWidget(self.__buttonChoix,1,1)
        layout1.addWidget(self.__buttonChoixLabel,2,1)
        layout1.addWidget(self.__nul2, 4, 1)
        layout1.addWidget(self.__error1, 0, 5)
        layout1.addWidget(self.__error11, 1, 5)
        layout1.addWidget(self.__labelText, 0, 1)

        self.__labelFanuc = QtWidgets.QLabel("Selezionare il tipo di file")

        self.boutonFanuc = QtWidgets.QPushButton("Fanuc")
        self.boutonSchlong = QtWidgets.QPushButton("She hong")
        layout1.addWidget(self.boutonFanuc, 5, 2)
        layout1.addWidget(self.boutonSchlong, 5, 0)

        self.widget1 = QtWidgets.QWidget()
        self.widget1.setLayout(layout1)
        self.setCentralWidget(self.widget1)

        ##layout2 : avortée
        '''
        layout2 = QtWidgets.QGridLayout()
        self.__number = QtWidgets.QLabel('1')
        self.__trou = QtWidgets.QLabel('foro :')
        self.__coord = QtWidgets.QLabel('')
        self.__coord2 = QtWidgets.QLabel('coordinamento : ')
        self.boutonLire2 = QtWidgets.QPushButton("entrata")
        self.__champProfondeur = QtWidgets.QLineEdit("")
        self.__champProfondeur.setPlaceholderText("5")
        self.__labelProfondeur = QtWidgets.QLabel("Profondità (mm) :")
        self.__champVitesse = QtWidgets.QLineEdit("")
        self.__champVitesse.setPlaceholderText("5")
        self.__labelVitesse = QtWidgets.QLabel("Velocità (mm/s) :")
        self.__champRetrait = QtWidgets.QLineEdit("")
        self.__champRetrait.setPlaceholderText("5")
        self.__labelRetrait = QtWidgets.QLabel("Profondità (mm) :")
        layout2.addWidget(self.__champProfondeur, 1, 1)
        layout2.addWidget(self.__labelProfondeur,1,0)
        layout2.addWidget(self.__champVitesse,2,1)
        layout2.addWidget(self.__labelVitesse,2,0)
        layout2.addWidget(self.__champRetrait, 3, 1)
        layout2.addWidget(self.__labelRetrait, 3, 0)
        layout2.addWidget(self.boutonLire2, 4, 0)
        layout2.addWidget(self.__number, 0, 1)
        layout2.addWidget(self.__trou, 0, 0)
        layout2.addWidget(self.__coord2, 0, 3)
        layout2.addWidget(self.__coord, 0, 4)
        layout2.addWidget(self.__error2, 4, 4)
        self.widget2=QtWidgets.QWidget()
        self.widget2.setLayout(layout2)
        '''

        ##layout3 : G et type de troue

        layout3 = QtWidgets.QGridLayout()
        self.__GxxInput = QtWidgets.QLineEdit('G54')
        self.__GxxInput.setPlaceholderText("G54")
        self.__error3 = QtWidgets.QLabel('')

        self.boutonForaturia = QtWidgets.QPushButton("Foraturia")
        self.boutonFillettatura = QtWidgets.QPushButton("Filettatura")
        self.boutonAlesaggio = QtWidgets.QPushButton("Alesaggio")
        self.__interval = QtWidgets.QLabel("Sistema di coordinate")
        layout3.addWidget(self.boutonForaturia, 1, 0)
        layout3.addWidget(self.boutonFillettatura, 1, 1)
        layout3.addWidget(self.boutonAlesaggio, 1, 2)
        layout3.addWidget(self.__GxxInput, 0, 1)
        layout3.addWidget(self.__interval, 0, 0)
        layout3.addWidget(self.__error3, 0, 2)

        self.widget3 = QtWidgets.QWidget()
        self.widget3.setLayout(layout3)

        ##layout 4 : pour les foraturia
        try:
            fileDir = os.getcwd()
            fileExt = r"*.xlsx"
            dir=list(pathlib.Path(fileDir).glob(fileExt))[0]
            wb = openpyxl.load_workbook(str(dir))
            sheet1 = wb.active
        except:

            try:
                fileDir = os.getcwd()
                fileExt = r"*.xls"
                dir = list(pathlib.Path(fileDir).glob(fileExt))[0]
                wb = openpyxl.load_workbook(str(dir))
                sheet1 = wb.active

            except:
                self.__nul.setText('excel della tabella dei diametri non esiste (.xls o .xlsx)')
                return

        layout4 = QtWidgets.QGridLayout()
        self.__securite = QtWidgets.QLabel('z di sicurezza')
        layout4.addWidget(self.__securite, 3, 0)
        self.__champSecurite4 = QtWidgets.QLineEdit("")
        self.__champSecurite4.setPlaceholderText("100")
        layout4.addWidget(self.__champSecurite4, 3, 1)
        self.__error4 = QtWidgets.QLabel()
        layout4.addWidget(self.__error4, 0, 5)
        self.boutonEntrata4 = QtWidgets.QPushButton("entrata")
        layout4.addWidget(self.boutonEntrata4, 10, 1)
        self.__SxxInput = QtWidgets.QLineEdit("")
        self.__SxxInput.setPlaceholderText("650")
        layout4.addWidget(self.__SxxInput, 8, 1)
        self.__labelSxx = QtWidgets.QLabel("Speed(S) (t/s)")
        layout4.addWidget(self.__labelSxx, 8, 0)
        self.__FxxInput = QtWidgets.QLineEdit("")
        self.__FxxInput.setPlaceholderText("75")
        layout4.addWidget(self.__FxxInput, 9, 1)
        self.__labelFxx = QtWidgets.QLabel("Avanzamento(F) (mm)")
        layout4.addWidget(self.__labelFxx, 9, 0)
        self.__labelProfondeur4 = QtWidgets.QLabel("Profondità (mm) :")
        self.__champProfondeur4 = QtWidgets.QLineEdit("")
        self.__champProfondeur4.setPlaceholderText("30")
        layout4.addWidget(self.__labelProfondeur4, 4, 0)
        layout4.addWidget(self.__champProfondeur4, 4, 1)
        self.__R4 = QtWidgets.QLabel("quota di avvicimento (R) (mm) :")
        self.__champR4 = QtWidgets.QLineEdit("")
        self.__champR4.setPlaceholderText("2")
        layout4.addWidget(self.__champR4, 6, 1)
        layout4.addWidget(self.__R4, 6, 0)
        self.__Q4 = QtWidgets.QLabel("Q (mm) :")
        self.__champQ4 = QtWidgets.QLineEdit("")
        self.__champQ4.setPlaceholderText("2")
        layout4.addWidget(self.__Q4, 7, 0)
        layout4.addWidget(self.__champQ4, 7, 1)

        self.__diam4 = QtWidgets.QLabel("Diametro d'il foro (mm) :")
        self.__excelyes = QtWidgets.QLabel("Usa i valori in excel : ")
        self.__champdiam4 = QtWidgets.QLineEdit("")
        self.__champdiam4.setPlaceholderText("5")
        layout4.addWidget(self.__diam4, 1, 0)
        layout4.addWidget(self.__champdiam4, 1, 1)
        layout4.addWidget(self.__excelyes,0,0)

        #self.diam = QtWidgets.QComboBox()

        # RECHERCHE DES VALEURS DANS L'EXCEL
        '''  for row in sheet1.rows:
            print(row[0].value)
            d = str(row[0].value).split('-')
            print(d)
            d = d[-1]
            print(d)
            try:
                d=float(d)
                print('bonsoir')
                self.diam.addItem(str(row[0].value))
                print('mé')
            except:
                pass
        '''
        self.metal = QtWidgets.QComboBox()
        self.metal.addItem('Alluminio')
        self.metal.addItem('Acciaio')

        #layout4.addWidget(self.diam, 0, 1)
        layout4.addWidget(self.metal, 1, 2)

        self.diamo = QtWidgets.QCheckBox('Altro foro, senza usare valore del file excel')
        # self.cb.toggle()
        self.diamo.stateChanged.connect(self.newParameters2)
        layout4.addWidget(self.diamo, 0, 2)

        self.__labelFxx.hide()
        self.__FxxInput.hide()
        self.__labelSxx.hide()
        self.__SxxInput.hide()
        #self.__champdiam4.hide()
        #self.__diam4.hide()


        self.widget4 = QtWidgets.QWidget()
        self.widget4.setLayout(layout4)

        ##layout 5 : pour les fillatura
        layout5 = QtWidgets.QGridLayout()
        self.__securite5 = QtWidgets.QLabel('z di sicurezza')
        layout5.addWidget(self.__securite5, 1, 0)
        self.__champSecurite5 = QtWidgets.QLineEdit("")
        self.__champSecurite5.setPlaceholderText("100")
        layout5.addWidget(self.__champSecurite5, 1, 1)

        self.combo = QtWidgets.QComboBox()
        self.combo.addItem('M5')
        self.combo.addItem('M6')
        self.combo.addItem('M8')
        self.combo.addItem('M10')
        self.combo.addItem('M12')
        self.combo.addItem('M14')
        self.combo.addItem('M16')
        self.combo.addItem('M20')
        layout5.addWidget(self.combo, 0, 1)

        self.cb = QtWidgets.QCheckBox('Altro M')
        # self.cb.toggle()
        self.cb.stateChanged.connect(self.newParameters)
        layout5.addWidget(self.cb, 0, 0)

        self.__labelProfondeur = QtWidgets.QLabel("Profondità (mm) :")
        self.__champProfondeur = QtWidgets.QLineEdit("")
        self.__champProfondeur.setPlaceholderText("30")
        layout5.addWidget(self.__labelProfondeur, 4, 0)
        layout5.addWidget(self.__champProfondeur, 4, 1)
        self.__R = QtWidgets.QLabel("quota di avvicimento (mm) :")
        self.__champR = QtWidgets.QLineEdit("")
        self.__champR.setPlaceholderText("2")
        layout5.addWidget(self.__champR, 5, 1)
        layout5.addWidget(self.__R, 5, 0)
        self.__diam5 = QtWidgets.QLabel("Diametro d'il foro (mm) :")
        self.__champdiam5 = QtWidgets.QLineEdit("")
        self.__champdiam5.setPlaceholderText("5")
        layout5.addWidget(self.__diam5, 7, 0)
        layout5.addWidget(self.__champdiam5, 7, 1)
        self.__error5 = QtWidgets.QLabel()
        layout5.addWidget(self.__error5, 1, 5)

        self.boutonEntrata5 = QtWidgets.QPushButton("entrata")
        layout5.addWidget(self.boutonEntrata5, 8, 1)

        self.widget5 = QtWidgets.QWidget()
        self.widget5.setLayout(layout5)

        # params supplémentaires pour M personnalisé

        self.__labelNewSxx = QtWidgets.QLabel("Speed (t/m) :")
        self.__champNewSxx = QtWidgets.QLineEdit("")
        self.__champNewSxx.setPlaceholderText("650")

        self.__labelNewFxx = QtWidgets.QLabel("Avanzamento (F) (mm) :")
        self.__champNewFxx = QtWidgets.QLineEdit("")
        self.__champNewFxx.setPlaceholderText("75")

        layout5.addWidget(self.__labelNewSxx, 2, 0)
        layout5.addWidget(self.__champNewSxx, 2, 1)
        layout5.addWidget(self.__labelNewFxx, 3, 0)
        layout5.addWidget(self.__champNewFxx, 3, 1)

        self.__labelNewSxx.hide()
        self.__champNewSxx.hide()
        self.__labelNewFxx.hide()
        self.__champNewFxx.hide()

        ##layout6 : alesage
        layout6 = QtWidgets.QGridLayout()
        self.widget6 = QtWidgets.QWidget()
        self.widget6.setLayout(layout6)

        ##appel de fonction

        self.boutonFanuc.clicked.connect(self.fanuc)
        self.boutonSchlong.clicked.connect(self.read)
        self.__buttonChoix.clicked.connect(self.choix)
        self.boutonFillettatura.clicked.connect(self.fillettatura)
        self.boutonForaturia.clicked.connect(self.foraturia)
        self.boutonAlesaggio.clicked.connect(self.alesaggio)
        self.boutonEntrata5.clicked.connect(self.writefi)
        self.boutonEntrata4.clicked.connect(self.writefo)
        # self.boutonLire.clicked.connect(self.read)
        # self.boutonLire2.clicked.connect(self.FtoPayRespects)
        # self.boutonSecurite.clicked.connect(self.prof)

        wb.close()

    ##les fonctions

    def rechercheIntervalle(self,diametre):
        try:
            fileDir = os.getcwd()
            fileExt = r"*.xlsx"
            dir = list(pathlib.Path(fileDir).glob(fileExt))[0]
            wb = openpyxl.load_workbook(str(dir))
            sheet1 = wb.active
        except:

            try:
                fileDir = os.getcwd()
                fileExt = r"*.xls"
                dir = list(pathlib.Path(fileDir).glob(fileExt))[0]
                wb = openpyxl.load_workbook(str(dir))
                sheet1 = wb.active

            except:
                self.__error4.setText('excel della tabella dei diametri non esiste (.xls o .xlsx)')
                return

        for row in sheet1.rows:
               diam=str(row[0].value).split('-')
               try:
                   haut = float(diam[-1])
                   bas = float(diam[0])
                   if bas<=float(diametre)<haut:
                       if self.metal.currentText() == 'Alluminio':
                           self.__SxxInput.setText(str(row[1].value))
                           self.__FxxInput.setText(str(row[2].value))
                           wb.close()
                           return
                       else:
                           self.__SxxInput.setText(str(row[3].value))
                           self.__FxxInput.setText(str(row[4].value))
                           wb.close()
                           return
               except:
                   pass
        wb.close()


    def choix(self):
        root = tk.Tk()
        root.withdraw()  # pour ne pas afficher la fenêtre Tk

        self.chosenFile = askopenfilename()
        if not'.igs' in self.chosenFile:
            self.__nul.setText("il file non e .igs")
            self.chosenFile=''
            return
        self.__buttonChoixLabel.setText('File : '+self.chosenFile)
        self.__nul.clear()

    def testGxx(self):
        input = self.__GxxInput.text()
        test = re.findall("^G5[4-9]$", input)
        if test:
            self.Gxx = self.__GxxInput.text()
            return True
        else:
            self.__GxxInput.clear()
            self.__error3.setText('questo comando non esiste')
            return False

    def newParameters(self):
        if (self.combo.count() == 8):
            self.combo.addItem('Altro M')
            self.combo.setCurrentIndex(8)
            for i in range(8):
                self.combo.removeItem(0)
            self.__labelNewSxx.show()
            self.__champNewSxx.show()
            self.__labelNewFxx.show()
            self.__champNewFxx.show()

        else:
            self.combo.removeItem(0)
            self.combo.addItem('M5')
            self.combo.addItem('M6')
            self.combo.addItem('M8')
            self.combo.addItem('M10')
            self.combo.addItem('M12')
            self.combo.addItem('M14')
            self.combo.addItem('M16')
            self.combo.addItem('M20')
            self.__labelNewSxx.hide()
            self.__champNewSxx.hide()
            self.__labelNewFxx.hide()
            self.__champNewFxx.hide()

    def newParameters2(self):
       ''' if (self.diam.count() > 1):
            self.diam.addItem('Altro')
            self.diam.setCurrentIndex(self.diam.count() - 1)
            for diam in range(self.diam.count() - 1):
                self.diam.removeItem(0)
            self.__labelFxx.show()
            self.__FxxInput.show()
            self.__labelSxx.show()
            self.__SxxInput.show()
            self.metal.hide()
            self.__champdiam4.show()
            self.__diam4.show()

        else:
            try:
                fileDir = os.getcwd()
                fileExt = r"*.xlsx"
                dir = list(pathlib.Path(fileDir).glob(fileExt))[0]
                wb = openpyxl.load_workbook(dir)
                sheet1 = wb.active
            except:

                try:
                    fileDir = os.getcwd()
                    fileExt = r"*.xls"
                    dir = list(pathlib.Path(fileDir).glob(fileExt))[1]
                    wb = openpyxl.load_workbook(dir)
                    sheet1 = wb.active

                except:
                    self.__nul.setText('excel della tabella dei diametri non esiste (.xls o .xlsx)')
                    return
            self.diam.removeItem(0)
            for row in sheet1.rows:
                if type(row[0].value) == int or type(row[0].value) == float:
                    self.diam.addItem(str(row[0].value))
            self.__labelFxx.hide()
            self.__FxxInput.hide()
            self.__labelSxx.hide()
            self.__SxxInput.hide()
            self.metal.show()
            self.__champdiam4.hide()
            self.__diam4.hide()
            wb.close()
            '''
       if(self.diamo.isChecked()):
            self.__labelFxx.show()
            self.__FxxInput.show()
            self.__labelSxx.show()
            self.__SxxInput.show()
            self.metal.hide()
       else:
           self.__labelFxx.hide()
           self.__FxxInput.hide()
           self.__labelSxx.hide()
           self.__SxxInput.hide()

           self.__FxxInput.clear()
           self.__SxxInput.clear()

           self.metal.show()


    def foraturia(self):
        if not (self.testGxx()):
            return
        self.setCentralWidget(self.widget4)

    def fillettatura(self):
        if not self.testGxx():
            return
        self.setCentralWidget(self.widget5)

    def alesaggio(self):
        if not self.testGxx():
            return
        self.setCentralWidget(self.widget6)

    def fanuc(self):
        self.Fanuc = True
        self.read()

    def read(self):
        self.__error1.clear()
        filename = self.chosenFile

        try:
            with open(filename, 'r') as f:

                param_string = ''
                entity_index = 0
                first_dict_line = True
                first_global_line = True
                first_param_line = True
                global_string = ""
                pointer_dict = {}

                for line in f.readlines():
                    data = line[:80]
                    id_code = line[72]
                    if id_code == 'S':  # Start
                        desc = line[:72].strip()

                    elif id_code == 'G':  # Global
                        global_string += data  # Consolidate all global lines
                        if first_global_line:
                            param_sep = data[2]
                            record_sep = data[6]
                            first_global_line = False

                    elif id_code == 'D':  # Directory entry
                        if first_dict_line:
                            entity_type_number = int(data[0:8].strip())
                            if entity_type_number == 116:  # Point
                                e = Point()
                                e.add_section(data[0:8], 'entity_type_number')
                                e.add_section(data[8:16], 'parameter_pointer')
                                e.add_section(data[16:24], 'structure')
                                e.add_section(data[24:32], 'line_font_pattern')
                                e.add_section(data[32:40], 'level')
                                e.add_section(data[40:48], 'view')
                                e.add_section(data[48:56], 'transform')
                                e.add_section(data[56:65], 'label_assoc')
                                e.add_section(data[65:72], 'status_number')
                                e.sequence_number = int(data[73:].strip())

                            first_dict_line = False

                        else:
                            if entity_type_number == 116:
                                e.add_section(data[8:16], 'line_weight_number')
                                e.add_section(data[16:24], 'color_number')
                                e.add_section(data[24:32], 'param_line_count')
                                e.add_section(data[32:40], 'form_number')
                                e.add_section(data[56:64], 'entity_label', type='string')
                                e.add_section(data[64:72], 'entity_subs_num')

                                self.points.append(e)
                                pointer_dict.update({e.sequence_number: entity_index})
                                entity_index += 1

                            first_dict_line = True

                    elif id_code == 'P':  # Parameter data
                        for x in pointer_dict:
                            print()
                            # Concatenate multiple lines into one string
                        if first_param_line:
                            param_string = data[:64]
                            directory_pointer = int(data[64:72].strip())
                            first_param_line = False
                        else:
                            param_string += data[:64]

                        if param_string.strip()[-1] == record_sep:
                            first_param_line = True
                            param_string = param_string.strip()[:-1]
                            parameters = param_string.split(param_sep)
                            if(parameters[0]=='116'):
                                 pt = self.points[pointer_dict[directory_pointer]]
                                 pt._add_parameters(parameters)
                                 zref = self.points[0].coordinate[2]
                                 if pt.coordinate[2] != zref:
                                     self.__error11.setText('Tutti i punti non hanno la stessa coordinata in z')
                                     return

                    elif id_code == 'T':  # Terminate
                        for e in self.points:
                            print()

            self.setCentralWidget(self.widget3)
            # self.__coord.setText(str(self.points[0].coordinate))
        except FileNotFoundError:
            self.__error1.setText('Il file non esiste')
            self.__champTexte.clear()
            return

    def FtoPayRespects(self):

        if self.compteur < len(self.points):
            self.compteur += 1
            info = self.__champProfondeur.text()
            info2 = self.__champVitesse.text()
            info3 = self.__champRetrait.text()
            if ',' in info or ',' in info2 or ',' in info3:
                try:
                    info = float(info.replace(',', '.'))
                    info2 = float(info2.replace(',', '.'))
                    info3 = float(info3.replace(',', '.'))
                except ValueError:
                    self.__error2.setText('input deve essere un valore numerico')
                    self.__champProfondeur.clear()
                    self.__champVitesse.clear()
                    self.__champRetrait.clear()
                    self.compteur -= 1
                    return
            try:
                info = float(info)
                info2 = float(info2)
                info3 = float(info3)
            except ValueError:
                self.__error2.setText('input deve essere un valore numerico')
                self.__champProfondeur.clear()
                self.__champVitesse.clear()
                self.__champRetrait.clear()
                self.compteur -= 1
                return
            self.__champProfondeur.clear()
            self.__champVitesse.clear()
            self.__champRetrait.clear()
            self.infopoints.append([info, info2, info3])
            if self.compteur < len(self.points):
                self.__number.setText(str(self.compteur + 1))
                self.__coord.setText(str(self.points[self.compteur].coordinate))
            if self.compteur == len(self.points):
                self.setCentralWidget(self.widget3)

    def write(self, option):
        if self.Fanuc:
            new = open('O0001', 'w')
            new.write('%\nO0001\n')
        else:
            new = open('1.PRG', 'w')
        new.write(self.Gxx + '\n')
        new.write(self.Sxx + 'M3\n')
        new.write('G0 ' + self.Securite + '\n')
        new.write('G98\n')
        for point in self.points:
            if option == 'fi':
                new.write('G84')
            elif option == 'fo':
                new.write('G83')
            else:
                new.write('G85')
            new.write('X' + str(round(float(point.x),3)) + 'Y' + str(round(float(point.y),3)) + self.Profondeur + self.Rxx + self.Qxx + self.Fxx + '\n')
        new.write('M30')
        if (self.Fanuc):
            new.write('\n%')

    def writeal(self):
        option = 'al'
        pass

    def writefo(self):
        option = 'fo'
        self.__error4.clear()

        '''try:
            wb = openpyxl.load_workbook('3.xlsx')
            sheet1 = wb.active
        except:
            print('non')
            return
        if self.diam.currentText() == 'Altro':
            pass
        else:
            for row in sheet1.rows:
                diam=str(row[0].value).split('-')
                diam=diam[-1]
                if str(str(row[0].value)) == self.diam.currentText():
                    self.__champdiam4.setText(diam)
                    if self.metal.currentText() == 'Alluminio':
                        self.__SxxInput.setText(str(row[1].value))
                        self.__FxxInput.setText(str(row[2].value))
                    else:
                        self.__SxxInput.setText(str(row[3].value))
                        self.__FxxInput.setText(str(row[4].value))
        '''
        diametre = self.__champdiam4.text()
        self.rechercheIntervalle(diametre)


        security = self.__champSecurite4.text()
        speed = self.__SxxInput.text()
        avanzamento = self.__FxxInput.text()
        profondeur = self.__champProfondeur4.text()
        r = self.__champR4.text()
        q = self.__champQ4.text()
        verif = [security, speed, avanzamento, profondeur, r, q]

        print(verif)

        for point1 in self.points:
            compt = 0
            for point2 in self.points:
                if ((float(point1.coordinate[0])-float(point2.coordinate[0]))**2+(float(point1.coordinate[1])-float(point2.coordinate[1]))**2)**1/2<float(diametre)+1:
                    compt+=1
            if compt>1:
                self.__error4.setText('due punti o due fori si sovrappongono')
                return



        for element in verif:
            if element == security:
                if ',' in element:
                    try:
                        element = float(element.replace(',', '.'))
                    except ValueError:
                        self.__error4.setText('input deve essere un valore numerico')
                        self.__champSecurite.clear()
                        return
                try:
                    element = float(element)
                except ValueError:
                    self.__error4.setText('input deve essere un valore numerico')
                    self.__champSecurite.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error4.setText('intervallo sbagliato per la sicurezza')
                    self.__champSecurite.clear()
                    return
            if element == profondeur:
                if ',' in element:
                    try:
                        element = float(element.replace(',', '.'))
                    except ValueError:
                        self.__error4.setText('input deve essere un valore numerico')
                        self.__champProfondeur4.clear()
                        return
                try:
                    element = float(element)
                except ValueError:
                    self.__error4.setText('input deve essere un valore numerico')
                    self.__champProfondeur4.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error4.setText('intervallo sbagliato per la profondità')
                    self.__champProfondeur4.clear()
                    return
            if element == r:
                if ',' in element:
                    try:
                        element = float(element.replace(',', '.'))
                    except ValueError:
                        self.__error4.setText('input deve essere un valore numerico')
                        self.__champR4.clear()
                        return
                try:
                    element = float(element)
                except ValueError:
                    self.__error4.setText('input deve essere un valore numerico')
                    self.__champR4.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error4.setText('intervallo sbagliato per la quota di avvicimento')
                    self.__champR4.clear()
                    return

            if element == q:
                if ',' in element:
                    try:
                        element = float(element.replace(',', '.'))
                    except ValueError:
                        self.__error4.setText('input deve essere un valore numerico')
                        self.__champQ4.clear()
                        return
                try:
                    element = float(element)
                except ValueError:
                    self.__error4.setText('input deve essere un valore numerico')
                    self.__champQ4.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error4.setText('intervallo sbagliato per Q')
                    self.__champQ4.clear()
                    return

            if element == speed:
                try:
                    element = int(element)
                except ValueError:
                    self.__error4.setText('input deve essere un valore numerico')
                    self.__champQ4.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error4.setText('intervallo sbagliato per la velocità')
                    self.__champQ4.clear()
                    return

            if element == avanzamento:
                try:
                    element = int(element)
                except ValueError:
                    self.__error4.setText('input deve essere un valore numerico')
                    self.__champQ4.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error4.setText('intervallo sbagliato per la avanzamento')
                    self.__champQ4.clear()
                    return

        self.Securite = 'Z' + str(security)
        self.Sxx = 'S' + str(speed)
        self.Fxx = 'F' + str(avanzamento)
        self.Profondeur = 'Z-' + str(profondeur)
        self.Qxx = 'Q' + str(q)
        self.Rxx = 'R' + str(r)

        try:
            self.write(option)
        except:
            return

        self.setCentralWidget(self.widget7)



    def writefi(self):
        option = 'fi'
        self.__error5.clear()
        self.Mx = self.combo.currentText()

        if self.Mx == 'M5':
            self.Sxx = 'S50'
            self.Fxx = 'F40'
        elif self.Mx == 'M6':
            self.Sxx = 'S50'
            self.Fxx = 'F50'
        elif self.Mx == 'M8':
            self.Sxx = 'S60'
            self.Fxx = 'F75'
        elif self.Mx == 'M10':
            self.Sxx = 'S50'
            self.Fxx = 'F75'
        elif self.Mx == 'M12':
            self.Sxx = 'S40'
            self.Fxx = 'F70'
        elif self.Mx == 'M14' or self.Mx == 'M16':
            self.Sxx = 'S50'
            self.Fxx = 'F100'
        elif self.Mx == 'M20':
            self.Sxx = 'S50'
            self.Fxx = 'F125'
        else:
            CurrSxx = self.__champNewSxx.text()
            CurrFxx = self.__champNewFxx.text()
            try:
                CurrSxx = int(CurrSxx)
            except ValueError:
                self.__error4.setText('input deve essere un valore numerico')
                self.__champQ4.clear()
                return
            if CurrSxx < 0 or CurrSxx > 10000:
                self.__error4.setText('intervallo sbagliato per la velocità')
                self.__champQ4.clear()
                return

            try:
                CurrFxx = int(CurrFxx)
            except ValueError:
                self.__error4.setText('input deve essere un valore numerico')
                self.__champQ4.clear()
                return
            if CurrFxx < 0 or CurrFxx > 10000:
                self.__error4.setText('intervallo sbagliato per la avanzamento')
                self.__champQ4.clear()
                return

            self.Sxx = 'S' + str(CurrSxx)
            self.Fxx = 'F' + str(CurrFxx)

        ##verification
        security = self.__champSecurite5.text()
        profondeur = self.__champProfondeur.text()
        r = self.__champR.text()
        diametre=self.__champdiam5.text()

        for point1 in self.points:
            compt = 0
            for point2 in self.points:
                if ((float(point1.coordinate[0])-float(point2.coordinate[0]))**2+(float(point1.coordinate[1])-float(point2.coordinate[1]))**2)**1/2<float(diametre)+1:
                    compt+=1
            if compt>1:
                self.__error5.setText('due punti o due fori si sovrappongono')
                return

        verif = [security, profondeur, r]
        for element in verif:
            if element == security:
                if ',' in element:
                    try:
                        element = float(element.replace(',', '.'))
                    except ValueError:
                        self.__error5.setText('input deve essere un valore numerico')
                        self.__champSecurite5.clear()
                        return
                try:
                    element = float(element)
                except ValueError:
                    self.__error5.setText('input deve essere un valore numerico')
                    self.__champSecurite5.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error5.setText('intervallo sbagliato per la sicurezza')
                    self.__champSecurite5.clear()
                    return
            if element == profondeur:
                if ',' in element:
                    try:
                        element = float(element.replace(',', '.'))
                    except ValueError:
                        self.__error5.setText('input deve essere un valore numerico')
                        self.__champProfondeur.clear()
                        return
                try:
                    element = float(element)
                except ValueError:
                    self.__error5.setText('input deve essere un valore numerico')
                    self.__champProfondeur.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error5.setText('intervallo sbagliato per la profondità')
                    self.__champProfondeur.clear()
                    return
            if element == r:
                if ',' in element:
                    try:
                        element = float(element.replace(',', '.'))
                    except ValueError:
                        self.__error5.setText('input deve essere un valore numerico')
                        self.__champR.clear()
                        return
                try:
                    element = float(element)
                except ValueError:
                    self.__error5.setText('input deve essere un valore numerico')
                    self.__champR.clear()
                    return
                if element < 0 or element > 10000:
                    self.__error5.setText('intervallo sbagliato per la quota di avvicimento')
                    self.__champR.clear()
                    return

        self.Securite = 'Z' + str(security)
        self.Profondeur = 'Z-' + str(profondeur)
        self.Qxx = ''
        self.Rxx = 'R' + str(r)
        try:
            self.write(option)
        except:
            return

        self.setCentralWidget(self.widget7)


class Point():
    def __init__(self):
        self.d = dict()
        self.parameters = []

    def add_section(self, string, key, type='int'):
        string = string.strip()
        if type == 'string':
            self.d[key] = string
        else:
            # Status numbers are made of four 2-digit numbers, together making an 8-digit number.
            # It includes spaces, so  0 0 0 0 is a valid 8-digit for which int casting won't work.
            if key == "status_number":
                # Get a list of four 2-digit numbers with spaces removed.
                separated_status_numbers = [string[i:i + 2].replace(' ', '0') for i in range(0, len(string), 2)]

                # Join these status numbers together as a single string.
                status_number_string = ''.join(separated_status_numbers)

                # The string can now be properly cast as an int
                self.d[key] = int(status_number_string)
            elif len(string) > 0:
                self.d[key] = int(string)
            else:
                self.d[key] = None

    def _add_parameters(self, parameters):
        self._x = parse_float(parameters[1])
        self._y = parse_float(parameters[2])
        self._z = parse_float(parameters[3])

    @property
    def x(self):
        """X coordinate"""
        return self._x

    @property
    def y(self):
        """Y coordinate"""
        return self._y

    @property
    def z(self):
        """Z coordinate"""
        return self._z

    @property
    def coordinate(self):
        """Coordinate of the point as a numpy array"""
        return np.array([self._x, self._y, self._z])


app = QtWidgets.QApplication(sys.argv)

window = MaFenetre()
window.show()

app.exec_()